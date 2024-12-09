from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def apply_excel_formatting(file_path, column_to_format, pass_value="Pass", fail_value="Fail"):
    """
    Applies formatting to the Excel file: Green for 'Pass', Red for 'Fail'.
    
    Parameters:
        file_path (str): Path to the Excel file to format.
        column_to_format (str): Name of the column to apply formatting to.
        pass_value (str): Value in the column considered as "Pass".
        fail_value (str): Value in the column considered as "Fail".
    """
    workbook = load_workbook(file_path)
    sheet = workbook.active

    # Define fills for Pass and Fail
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Find the column index of the specified column
    header = [cell.value for cell in sheet[1]]
    if column_to_format not in header:
        raise ValueError(f"Column '{column_to_format}' not found in the Excel file.")
    column_index = header.index(column_to_format) + 1  # Convert to 1-based index

    # Apply formatting
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=column_index, max_col=column_index):
        for cell in row:
            if cell.value == pass_value:
                cell.fill = green_fill
            elif cell.value == fail_value:
                cell.fill = red_fill

    # Save the formatted workbook
    workbook.save(file_path)
    print(f"Formatting applied and saved to '{file_path}'")
